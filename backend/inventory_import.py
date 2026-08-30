"""Importador inteligente de inventario (CSV / Excel) con detección flexible de columnas."""

import csv
import io
import os
import re
import traceback
import unicodedata

import openpyxl

from config import MAX_UPLOAD_BYTES
from backend.image_lookup import EXPR_CODIGO_BARRAS
from backend.utils import (
    imagen_url_almacenada,
    imagen_url_para_persistir,
    normalizar_codigo_barras,
)

SINONIMOS_COLUMNA = {
    'nombre': [
        'nombre',
        'producto',
        'articulo',
        'artículo',
        'descripcion',
        'descripción',
        'item',
        'denominacion',
        'denominación',
        'concepto',
    ],
    'precio': [
        'precio',
        'costo',
        'pvp',
        'precio usd',
        'precio_usd',
        'precio bs',
        'precio_bs',
        'precio venta',
        'valor',
        'monto',
        'importe',
        'precio unitario',
    ],
    'stock': [
        'cantidad',
        'stock',
        'existencia',
        'existencias',
        'inventario',
        'qty',
        'unidades',
        'disponible',
    ],
    'descripcion': [
        'detalle',
        'observaciones',
        'notas',
        'descripcion larga',
        'descripción larga',
        'info adicional',
    ],
    'codigo_barras': [
        'codigo barras',
        'codigo_barras',
        'codigo de barras',
        'barcode',
        'ean',
        'sku',
        'codigo',
        'referencia',
    ],
    'imagen_url': [
        'imagen url',
        'imagen_url',
        'url imagen',
        'imagen',
        'foto',
        'fotos',
        'image',
        'photo',
        'picture',
        'img',
        'link imagen',
        'url foto',
        'foto url',
    ],
}

ETIQUETAS_COLUMNA = {
    'nombre': 'Nombre del producto (nombre, producto, artículo, descripción…)',
    'precio': 'Precio (precio, costo, PVP, precio_usd, precio_bs…)',
    'stock': 'Cantidad / stock (cantidad, stock, existencia, inventario…)',
}

CAMPOS_OBLIGATORIOS = ('nombre', 'precio')
_MAX_ERRORES_VALIDACION = 20
_UMBRAL_COINCIDENCIA = 50

MAX_IMPORT_FILE_BYTES = min(
    int(os.getenv('MAX_IMPORT_FILE_BYTES', str(MAX_UPLOAD_BYTES))),
    MAX_UPLOAD_BYTES,
)
IMPORT_BATCH_SIZE = int(os.getenv('IMPORT_BATCH_SIZE', '500'))
LOG_PREFIX = '[Localis CSV]'
_MAX_FLASH_CHARS = 1400

INSERT_PRODUCTO_SQL = """
    INSERT INTO productos (
        comercio_id, nombre, descripcion, precio_usd,
        codigo_barras, imagen_url, stock
    )
    VALUES (?, ?, ?, ?, ?, ?, ?)
"""


class ErrorImportacionInventario(Exception):
    """Error controlado durante importación masiva (mensaje amigable para el comercio)."""


def normalizar_encabezado(texto):
    if texto is None:
        return ''
    normalizado = unicodedata.normalize('NFKD', str(texto))
    normalizado = ''.join(c for c in normalizado if not unicodedata.combining(c))
    normalizado = normalizado.lower().strip()
    normalizado = re.sub(r'[_\-./\\]+', ' ', normalizado)
    normalizado = re.sub(r'\s+', ' ', normalizado)
    return normalizado


def _puntuacion_coincidencia(encabezado_norm, sinonimo):
    if not encabezado_norm or not sinonimo:
        return 0
    if encabezado_norm == sinonimo:
        return 100
    if encabezado_norm.startswith(f'{sinonimo} '):
        return 85
    palabras = encabezado_norm.split()
    if sinonimo in palabras:
        return 75
    if encabezado_norm.endswith(f' {sinonimo}'):
        return 70
    if sinonimo in encabezado_norm:
        return 55
    return 0


def detectar_mapeo_columnas(encabezados):
    """
    Detecta columnas por sinónimos en la primera fila.
    Retorna (mapeo, meta, error).
    """
    encabezados_originales = [
        str(h).strip()
        for h in (encabezados or [])
        if h is not None and str(h).strip()
    ]
    if not encabezados_originales:
        return None, None, 'El archivo no tiene encabezados en la primera fila.'

    encabezados_norm = {
        orig: normalizar_encabezado(orig) for orig in encabezados_originales
    }
    usados = set()
    mapeo = {}
    meta = {'precio_en_bs': False}

    orden_campos = (
        'nombre',
        'precio',
        'stock',
        'descripcion',
        'codigo_barras',
        'imagen_url',
    )

    for campo in orden_campos:
        mejor_encabezado = None
        mejor_puntaje = 0

        for original, normalizado in encabezados_norm.items():
            if original in usados:
                continue
            for sinonimo in SINONIMOS_COLUMNA.get(campo, []):
                puntaje = _puntuacion_coincidencia(normalizado, sinonimo)
                if puntaje > mejor_puntaje:
                    mejor_puntaje = puntaje
                    mejor_encabezado = original

        if mejor_puntaje >= _UMBRAL_COINCIDENCIA and mejor_encabezado:
            mapeo[campo] = mejor_encabezado
            usados.add(mejor_encabezado)
            if campo == 'precio':
                norm_precio = encabezados_norm[mejor_encabezado]
                if any(
                    token in norm_precio
                    for token in ('bs', 'bolivar', 'bolívares', 'ves')
                ):
                    meta['precio_en_bs'] = True

    faltantes = [
        ETIQUETAS_COLUMNA[campo]
        for campo in CAMPOS_OBLIGATORIOS
        if campo not in mapeo
    ]
    if faltantes:
        detectadas = ', '.join(f'«{h}»' for h in encabezados_originales)
        return (
            None,
            None,
            'No pudimos reconocer columnas obligatorias en la primera fila. '
            f'Falta: {"; ".join(faltantes)}. '
            f'Columnas encontradas en tu archivo: {detectadas}. '
            'Corrige los encabezados e intenta de nuevo.',
        )

    return mapeo, meta, None


def _extension_archivo(archivo):
    if not archivo or not getattr(archivo, 'filename', ''):
        return None
    return archivo.filename.rsplit('.', 1)[-1].lower()


def recortar_mensaje_importacion(mensaje):
    """Evita flashes enormes que rompen la cookie de sesión (500 al redirigir)."""
    texto = str(mensaje or '').strip() or 'No se pudo completar la importación.'
    if len(texto) <= _MAX_FLASH_CHARS:
        return texto
    sufijo = '\n… (mensaje recortado).'
    return texto[: max(0, _MAX_FLASH_CHARS - len(sufijo))].rstrip() + sufijo


def _registrar_fallo_importacion(etapa, exc):
    print(f'{LOG_PREFIX} FALLO etapa={etapa} {type(exc).__name__}: {exc}')
    traceback.print_exc()


def cargar_archivo_inventario(archivo):
    """Lee el archivo subido con límite de tamaño. Retorna (bytes, extension, error)."""
    if not archivo or not getattr(archivo, 'filename', ''):
        return None, None, 'No se adjuntó ningún archivo.'

    extension = _extension_archivo(archivo)
    if extension not in {'csv', 'xlsx'}:
        return None, None, 'El archivo debe tener extensión .csv o .xlsx.'

    try:
        stream = getattr(archivo, 'stream', archivo)
        if hasattr(stream, 'seek'):
            stream.seek(0)

        data = stream.read(MAX_IMPORT_FILE_BYTES + 1)
    except Exception as exc:
        _registrar_fallo_importacion('leer_stream', exc)
        return None, None, 'No se pudo leer el archivo subido. Intenta de nuevo.'

    if not data:
        return None, None, 'El archivo está vacío.'

    if len(data) > MAX_IMPORT_FILE_BYTES:
        max_mb = MAX_IMPORT_FILE_BYTES // (1024 * 1024)
        return (
            None,
            None,
            f'El archivo supera el tamaño máximo permitido ({max_mb} MB). '
            'Divide el inventario en archivos más pequeños e intenta de nuevo.',
        )

    es_utf16 = data.startswith(b'\xff\xfe') or data.startswith(b'\xfe\xff')
    if extension == 'csv' and not es_utf16 and b'\x00' in data[:65536]:
        return (
            None,
            None,
            'El archivo no es un CSV de texto válido (parece binario o Excel). '
            'En Excel usa «Guardar como → CSV UTF-8» e intenta de nuevo.',
        )

    return data, extension, None


def _detectar_delimitador_csv(muestra):
    texto = muestra or ''
    primera_linea = next((linea for linea in texto.splitlines() if linea.strip()), '')
    if not primera_linea:
        return ','
    try:
        if len(primera_linea) >= 8 and (',' in primera_linea or ';' in primera_linea):
            dialecto = csv.Sniffer().sniff(primera_linea, delimiters=',;\t|')
            return dialecto.delimiter
    except (csv.Error, TypeError, ValueError):
        pass
    if primera_linea.count(';') >= primera_linea.count(',') and ';' in primera_linea:
        return ';'
    if '\t' in primera_linea:
        return '\t'
    if '|' in primera_linea:
        return '|'
    return ','


_ENCODINGS_CSV = ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1', 'iso-8859-1')
_MENSAJE_CODIFICACION_CSV = (
    'No se pudo leer la codificación del archivo CSV. '
    'Guárdalo como UTF-8 o Latin-1 (Excel: «CSV UTF-8») e intenta de nuevo.'
)


def _decodificar_csv(data):
    if not data:
        return None
    if data.startswith(b'\xff\xfe') or data.startswith(b'\xfe\xff'):
        try:
            texto = data.decode('utf-16')
            if '\x00' not in texto:
                return texto
        except UnicodeDecodeError:
            pass
    if b'\x00' in data[:65536]:
        return None
    for encoding in _ENCODINGS_CSV:
        try:
            texto = data.decode(encoding)
        except UnicodeDecodeError:
            continue
        if '\x00' in texto:
            continue
        return texto
    return None


def leer_encabezados_inventario(data, extension):
    """Lee solo la primera fila de encabezados. Retorna (encabezados, error)."""
    try:
        if extension == 'xlsx':
            wb = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True
            )
            try:
                hoja = wb.active
                primera = next(hoja.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not primera:
                    return None, 'El archivo Excel está vacío.'
                encabezados = [str(v or '').strip() for v in primera]
                while encabezados and not encabezados[-1]:
                    encabezados.pop()
                if not encabezados:
                    return None, 'El archivo Excel no tiene encabezados en la primera fila.'
                return encabezados, None
            finally:
                wb.close()

        contenido = _decodificar_csv(data)
        if contenido is None:
            return None, _MENSAJE_CODIFICACION_CSV

        delimitador = _detectar_delimitador_csv(contenido[:4096])
        reader = csv.reader(io.StringIO(contenido), delimiter=delimitador)
        primera = next(reader, None)
        if not primera:
            return None, 'El archivo CSV está vacío.'
        encabezados = [str(h or '').strip() for h in primera if str(h or '').strip()]
        if not encabezados:
            return None, 'El archivo CSV no tiene encabezados en la primera fila.'
        return encabezados, None
    except csv.Error as exc:
        _registrar_fallo_importacion('encabezados_csv', exc)
        return None, (
            'El CSV no se pudo interpretar. Revisa la primera fila '
            '(delimitadores, comillas y caracteres especiales) y guárdalo como CSV UTF-8.'
        )
    except Exception as exc:
        _registrar_fallo_importacion('leer_encabezados', exc)
        return None, (
            'No se pudieron leer los encabezados del archivo. '
            'Verifica que sea un CSV o Excel válido e intenta de nuevo.'
        )


def _fila_tiene_datos(fila_dict):
    return any(str(v or '').strip() for v in fila_dict.values())


def _celda_vacia(valor):
    if valor is None:
        return True
    texto = str(valor).strip()
    return not texto or texto.lower() in ('none', 'null', 'nan', 'n/a', '-')


def iter_filas_inventario_enumeradas(data, extension, encabezados):
    """Generador (numero_fila, fila_dict). Datos desde fila 2 (Excel) o equivalente CSV."""
    if extension == 'xlsx':
        wb = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
        try:
            hoja = wb.active
            numero_fila = 1
            for row in hoja.iter_rows(min_row=2):
                numero_fila += 1
                if not any(
                    _valor_celda_excel(celda) not in (None, '')
                    for celda in row
                ):
                    continue
                fila = {}
                for idx, encabezado in enumerate(encabezados):
                    if not encabezado:
                        continue
                    celda = row[idx] if idx < len(row) else None
                    fila[encabezado] = _valor_celda_excel(celda)
                if _fila_tiene_datos(fila):
                    yield numero_fila, fila
        finally:
            wb.close()
        return

    contenido = _decodificar_csv(data)
    if contenido is None:
        raise ErrorImportacionInventario(_MENSAJE_CODIFICACION_CSV)

    delimitador = _detectar_delimitador_csv(contenido[:4096])
    try:
        reader = csv.DictReader(io.StringIO(contenido), delimiter=delimitador)
        for indice, fila in enumerate(reader, start=2):
            fila_limpia = {
                str(clave).strip(): valor
                for clave, valor in fila.items()
                if clave is not None and str(clave).strip()
            }
            if _fila_tiene_datos(fila_limpia):
                yield indice, fila_limpia
    except csv.Error as exc:
        _registrar_fallo_importacion('parsear_csv', exc)
        raise ErrorImportacionInventario(
            'El CSV no se pudo interpretar. Revisa delimitadores, comillas '
            'y caracteres especiales. Guárdalo como CSV UTF-8 e intenta de nuevo.'
        ) from exc


def iter_filas_inventario(data, extension, encabezados):
    """Generador de filas {encabezado: valor} sin cargar todo el inventario en RAM."""
    for _, fila in iter_filas_inventario_enumeradas(data, extension, encabezados):
        yield fila


def _valor_celda_excel(celda):
    """Lee valor o hipervínculo de una celda Excel (las fotos suelen ir como HYPERLINK)."""
    if celda is None:
        return None
    hyper = getattr(celda, 'hyperlink', None)
    if hyper is not None:
        destino = getattr(hyper, 'target', None) or getattr(hyper, 'display', None)
        if destino:
            return str(destino)
    return celda.value


def _obtener_valor_celda(fila, columna):
    if columna is None:
        return None
    if columna in fila:
        return fila[columna]
    columna_lower = str(columna).lower()
    for clave, valor in fila.items():
        if str(clave).lower() == columna_lower:
            return valor
    return None


def _parsear_numero(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto or texto.lower() == 'none':
        return None

    limpio = re.sub(r'[^\d,.\-]', '', texto)
    if not limpio:
        return None

    if ',' in limpio and '.' in limpio:
        if limpio.rfind(',') > limpio.rfind('.'):
            limpio = limpio.replace('.', '').replace(',', '.')
        else:
            limpio = limpio.replace(',', '')
    elif ',' in limpio:
        partes = limpio.split(',')
        if len(partes[-1]) == 2:
            limpio = ''.join(partes[:-1]).replace('.', '') + '.' + partes[-1]
        else:
            limpio = limpio.replace(',', '.')

    try:
        return float(limpio)
    except ValueError:
        return None


def _parsear_entero(valor):
    numero = _parsear_numero(valor)
    if numero is None:
        return 0
    return max(0, int(round(numero)))


def diagnosticar_fila_obligatoria(fila, mapeo, meta, numero_fila, tasa_dolar=1.0):
    """Detecta celdas vacías o inválidas en campos obligatorios. Retorna lista de mensajes."""
    errores = []
    nombre_raw = _obtener_valor_celda(fila, mapeo.get('nombre'))
    if _celda_vacia(nombre_raw):
        errores.append(
            f'Fila {numero_fila}: el nombre del producto está vacío (campo obligatorio).'
        )

    precio_val = _obtener_valor_celda(fila, mapeo.get('precio'))
    if _celda_vacia(precio_val):
        errores.append(
            f'Fila {numero_fila}: el precio está vacío (campo obligatorio).'
        )
    else:
        precio_raw = _parsear_numero(precio_val)
        if precio_raw is None:
            errores.append(
                f'Fila {numero_fila}: el precio «{precio_val}» no es un número válido.'
            )
        elif precio_raw < 0:
            errores.append(
                f'Fila {numero_fila}: el precio no puede ser negativo.'
            )

    return errores


def validar_inventario_previo(
    data, extension, encabezados, mapeo, meta, tasa_dolar=1.0
):
    """
    Validación previa sin escribir en PostgreSQL.
    Retorna (valido, mensaje_error, meta_validacion).
    """
    errores = []
    filas_con_datos = 0
    filas_validas = 0

    try:
        for numero_fila, fila in iter_filas_inventario_enumeradas(
            data, extension, encabezados
        ):
            filas_con_datos += 1
            errores_fila = diagnosticar_fila_obligatoria(
                fila, mapeo, meta, numero_fila, tasa_dolar=tasa_dolar
            )
            if errores_fila:
                errores.extend(errores_fila)
                continue
            if parsear_fila_inventario(
                fila, mapeo, meta, tasa_dolar=tasa_dolar, imagen_default=None
            ):
                filas_validas += 1
    except ErrorImportacionInventario:
        raise
    except Exception as exc:
        _registrar_fallo_importacion('validacion_filas', exc)
        raise ErrorImportacionInventario(
            'No se pudieron leer las filas del archivo. '
            'Revisa la codificación y el formato e intenta de nuevo.'
        ) from exc

    if filas_con_datos == 0:
        return False, 'El archivo no contiene filas de datos debajo de los encabezados.', None

    if errores:
        visibles = errores[:_MAX_ERRORES_VALIDACION]
        mensaje = (
            'El archivo tiene errores en campos obligatorios. '
            'Ningún cambio fue aplicado a tu inventario.\n'
            + '\n'.join(visibles)
        )
        restantes = len(errores) - len(visibles)
        if restantes > 0:
            mensaje += f'\n… y {restantes} error(es) adicional(es).'
        return False, mensaje, {'errores': errores, 'filas_validas': filas_validas}

    if filas_validas == 0:
        return False, (
            'No se encontraron filas válidas con nombre y precio. '
            'Revisa que los datos no estén vacíos y que el precio use formato numérico.'
        ), None

    return True, None, {'filas_validas': filas_validas, 'filas_con_datos': filas_con_datos}


def parsear_fila_inventario(fila, mapeo, meta, tasa_dolar=1.0, imagen_default=None):
    """Convierte una fila del archivo en dict de producto o None si es inválida."""
    nombre_raw = _obtener_valor_celda(fila, mapeo.get('nombre'))
    nombre = str(nombre_raw or '').strip()
    if not nombre or nombre.lower() == 'none':
        return None

    precio_raw = _parsear_numero(_obtener_valor_celda(fila, mapeo.get('precio')))
    if precio_raw is None or precio_raw < 0:
        return None

    precio_usd = precio_raw
    if meta.get('precio_en_bs'):
        tasa = float(tasa_dolar or 1.0)
        if tasa <= 0:
            tasa = 1.0
        precio_usd = round(precio_raw / tasa, 2)
    else:
        precio_usd = round(precio_raw, 2)

    descripcion_col = mapeo.get('descripcion')
    descripcion = ''
    if descripcion_col:
        descripcion = str(_obtener_valor_celda(fila, descripcion_col) or '').strip()
        if descripcion.lower() == 'none':
            descripcion = ''

    codigo_barras = None
    if mapeo.get('codigo_barras'):
        codigo_barras = normalizar_codigo_barras(
            _obtener_valor_celda(fila, mapeo['codigo_barras'])
        )

    imagen_url = None
    if mapeo.get('imagen_url'):
        imagen_url = imagen_url_para_persistir(
            _obtener_valor_celda(fila, mapeo['imagen_url'])
        )

    stock = 0
    if mapeo.get('stock'):
        stock = _parsear_entero(_obtener_valor_celda(fila, mapeo['stock']))

    return {
        'nombre': nombre,
        'descripcion': descripcion,
        'precio_usd': precio_usd,
        'codigo_barras': codigo_barras,
        'imagen_url': imagen_url,
        'stock': stock,
    }


def contar_productos_validos(data, extension, encabezados, mapeo, meta, tasa_dolar=1.0, imagen_default=None):
    total = 0
    for fila in iter_filas_inventario(data, extension, encabezados):
        if parsear_fila_inventario(
            fila, mapeo, meta, tasa_dolar=tasa_dolar, imagen_default=imagen_default
        ):
            total += 1
    return total


def iter_lotes_productos(
    data,
    extension,
    encabezados,
    mapeo,
    meta,
    tasa_dolar=1.0,
    imagen_default=None,
    batch_size=None,
):
    """Genera lotes de productos parseados para inserción por bloques."""
    tamano = batch_size or IMPORT_BATCH_SIZE
    lote = []

    for fila in iter_filas_inventario(data, extension, encabezados):
        parsed = parsear_fila_inventario(
            fila, mapeo, meta, tasa_dolar=tasa_dolar, imagen_default=imagen_default
        )
        if not parsed:
            continue
        lote.append(parsed)
        if len(lote) >= tamano:
            yield lote
            lote = []

    if lote:
        yield lote


def _snapshot_imagenes_por_codigo(cursor, comercio_id):
    """
    Mapa codigo_barras normalizado → imagen_url existente antes de reemplazo masivo.
    Coincidencia estricta vía EXPR_CODIGO_BARRAS (PostgreSQL / Excel / CSV).
    """
    cursor.execute(
        f"""
        SELECT {EXPR_CODIGO_BARRAS} AS codigo_key, imagen_url
        FROM productos
        WHERE comercio_id = ?
          AND codigo_barras IS NOT NULL
          AND TRIM(BOTH FROM CAST(codigo_barras AS TEXT)) <> ''
        """,
        (int(comercio_id),),
    )
    snapshot = {}
    for fila in cursor.fetchall():
        if isinstance(fila, dict):
            clave_raw = fila.get('codigo_key')
            imagen_raw = fila.get('imagen_url')
        else:
            clave_raw, imagen_raw = fila[0], fila[1]
        clave = normalizar_codigo_barras(clave_raw)
        if not clave:
            continue
        imagen = imagen_url_almacenada(imagen_raw)
        if imagen:
            snapshot[clave] = imagen
    return snapshot


def _imagen_final_importacion(
    imagen_csv,
    codigo_barras,
    snapshot_imagenes,
    mapa_maestro=None,
    nombre=None,
    descripcion=None,
):
    """URL definitiva para INSERT: manual, snapshot local o image_manager."""
    del nombre, descripcion
    from backend.image_manager import resolver_imagen_escritura

    nueva = imagen_url_para_persistir(imagen_csv)
    if nueva:
        return nueva
    codigo = normalizar_codigo_barras(codigo_barras)
    if codigo and codigo in snapshot_imagenes:
        return snapshot_imagenes[codigo]
    return resolver_imagen_escritura(
        codigo_barras=codigo_barras,
        mapa_maestro=mapa_maestro,
    )


def _tuplas_insercion(comercio_id, lote, snapshot_imagenes=None, mapa_maestro=None):
    snapshot_imagenes = snapshot_imagenes or {}
    mapa_maestro = mapa_maestro or {}
    return [
        (
            comercio_id,
            prod['nombre'],
            prod['descripcion'],
            prod['precio_usd'],
            prod['codigo_barras'],
            _imagen_final_importacion(
                prod.get('imagen_url'),
                prod.get('codigo_barras'),
                snapshot_imagenes,
                mapa_maestro=mapa_maestro,
            ),
            prod['stock'],
        )
        for prod in lote
    ]


def persistir_importacion_por_lotes(comercio_id, factory_generador_lotes):
    """
    Transacción atómica con bloqueo por comercio e inserción por lotes.
    Las imágenes se resuelven antes del lock (catálogo maestro + OpenFoodFacts).
    """
    from backend.db import ejecutar_con_reintentos_bd, get_db_connection
    from backend.image_manager import preparar_mapa_imagenes_importacion

    productos_por_codigo = {}
    productos_sin_codigo = []
    for lote in factory_generador_lotes():
        for prod in lote:
            codigo = normalizar_codigo_barras(prod.get('codigo_barras'))
            if codigo:
                productos_por_codigo[codigo] = prod
            else:
                productos_sin_codigo.append(prod)

    productos_finales = list(productos_por_codigo.values()) + productos_sin_codigo

    with get_db_connection() as conexion:
        cursor = conexion.cursor()
        snapshot_imagenes = _snapshot_imagenes_por_codigo(cursor, comercio_id)

    mapa_imagenes = preparar_mapa_imagenes_importacion(
        productos_finales,
        snapshot_imagenes,
    )

    def _operacion(conexion):
        cursor = conexion.cursor()
        cursor.execute('SELECT pg_advisory_xact_lock(?)', (int(comercio_id),))
        cursor.execute('DELETE FROM productos WHERE comercio_id = ?', (int(comercio_id),))

        insertados = 0
        for inicio in range(0, len(productos_finales), IMPORT_BATCH_SIZE):
            lote = productos_finales[inicio : inicio + IMPORT_BATCH_SIZE]
            cursor.executemany(
                INSERT_PRODUCTO_SQL,
                _tuplas_insercion(comercio_id, lote, snapshot_imagenes, mapa_imagenes),
            )
            insertados += len(lote)

        if insertados == 0:
            raise ErrorImportacionInventario(
                'No se encontraron filas válidas con nombre y precio. '
                'Revisa que los datos no estén vacíos y que el precio use formato numérico.'
            )
        return insertados

    return ejecutar_con_reintentos_bd(_operacion)


def mensaje_error_importacion(exc):
    """Traduce excepciones técnicas a mensajes amigables para el comercio."""
    if isinstance(exc, ErrorImportacionInventario):
        return str(exc)
    if isinstance(exc, RuntimeError):
        return str(exc)

    exc_name = type(exc).__module__ + '.' + type(exc).__name__
    if exc_name in ('psycopg2.pool.PoolError', 'psycopg2.PoolError'):
        return (
            'El servidor está procesando muchas importaciones a la vez. '
            'Espera unos segundos e intenta de nuevo.'
        )
    if exc_name.startswith('psycopg2') and type(exc).__name__ == 'OperationalError':
        return (
            'Hubo una sobrecarga temporal al guardar tu inventario. '
            'Ningún cambio fue aplicado. Intenta de nuevo en unos segundos.'
        )
    try:
        from psycopg2 import OperationalError
        from psycopg2.pool import PoolError

        if isinstance(exc, PoolError):
            return (
                'El servidor está procesando muchas importaciones a la vez. '
                'Espera unos segundos e intenta de nuevo.'
            )
        if isinstance(exc, OperationalError):
            return (
                'Hubo una sobrecarga temporal al guardar tu inventario. '
                'Ningún cambio fue aplicado. Intenta de nuevo en unos segundos.'
            )
    except ImportError:
        pass

    if isinstance(exc, MemoryError):
        return (
            'El archivo es demasiado grande para procesarlo en este momento. '
            'Reduce el tamaño o divide el inventario en varios archivos.'
        )
    if isinstance(exc, csv.Error) or isinstance(exc, UnicodeError):
        return (
            'El CSV no se pudo leer. Revisa la codificación (UTF-8 o Latin-1), '
            'los delimitadores y que no haya caracteres binarios.'
        )
    return (
        'No se pudo completar la importación. '
        'Tus productos anteriores no fueron modificados. '
        'Verifica el archivo e intenta de nuevo.'
    )


# Compatibilidad con pruebas y llamadas legacy
def leer_filas_inventario(archivo):
    data, extension, error = cargar_archivo_inventario(archivo)
    if error:
        return None, None, error
    encabezados, error = leer_encabezados_inventario(data, extension)
    if error:
        return None, None, error
    filas = list(iter_filas_inventario(data, extension, encabezados))
    return filas, encabezados, None


def procesar_filas_inventario(filas, encabezados, tasa_dolar=1.0, imagen_default=None):
    mapeo, meta, error = detectar_mapeo_columnas(encabezados)
    if error:
        return None, error

    productos = []
    for fila in filas:
        parsed = parsear_fila_inventario(
            fila, mapeo, meta, tasa_dolar=tasa_dolar, imagen_default=imagen_default
        )
        if parsed:
            productos.append(parsed)

    if not productos:
        return None, (
            'No se encontraron filas válidas con nombre y precio. '
            'Revisa que los datos no estén vacíos y que el precio use formato numérico.'
        )

    return productos, None
